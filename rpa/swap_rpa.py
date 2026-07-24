# -*- coding: utf-8 -*-
"""
换图 RPA 执行脚本 - 在目标电脑由 swap_listener.py 调用
使用 Playwright 自动化拼多多商家后台换图操作

流程：
  1. 读取 Account.xlsx 获取店铺账号密码
  2. 根据目标商品所属店铺，逐个登录对应后台
  3. 下载源商品最佳图片
  4. 打开目标商品编辑页，替换最差图片并保存
  5. 记录结果

登录策略：自动填账号密码 → 暂停等待手动过滑块验证码 → 继续
"""
import json
import sys
import os
import time
import argparse
import subprocess
from datetime import datetime
from urllib.parse import urlparse

if os.name == "nt":
    import asyncio
    try:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    except Exception:
        pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS_DIR = os.path.join(SCRIPT_DIR, "swap_downloads")
os.makedirs(DOWNLOADS_DIR, exist_ok=True)

MMS_BASE = "https://mms.pinduoduo.com"
MMS_LOGIN = MMS_BASE + "/login"
GOODS_LIST = MMS_BASE + "/goods/goods_list"
GOODS_EDIT = MMS_BASE + "/goods/goods_edit"

# Account.xlsx 路径（与记忆中的路径一致）
ACCOUNT_FILE = r"C:\Users\s\Desktop\Result\Account.xlsx"
# 备用
if not os.path.exists(ACCOUNT_FILE):
    ACCOUNT_FILE = r"C:\Users\s\Desktop\Account.xlsx"


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = "[{}] {}".format(ts, msg)
    print(line, flush=True)


def sanitize_filename(url):
    parsed = urlparse(url)
    name = os.path.basename(parsed.path)
    if not name or len(name) < 5:
        name = "image_{}.jpg".format(int(time.time()))
    if "?" in name:
        name = name.split("?")[0]
    return name


def get_store_credentials():
    """从 Account.xlsx 读取店铺登录凭证
    返回: {store_name: {username, password}} 或空
    """
    creds = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ACCOUNT_FILE, data_only=True)
        if "拼多多" in wb.sheetnames:
            ws = wb["拼多多"]
        else:
            # 尝试第一个 sheet
            ws = wb.active
        log("读取凭据: {} (sheet: {})".format(ACCOUNT_FILE, ws.title))
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            store_name = str(row[0]).strip() if row[0] else ""
            username = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            password = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if store_name and username and password:
                creds[store_name] = {"username": username, "password": password}
                log("  凭据: {} -> {}".format(store_name, username))
        wb.close()
    except Exception as e:
        log("读取凭据失败: {}".format(e))
    return creds


def find_store_for_product(product_id):
    """根据 product_id 判断所属店铺（从数据文件中推测）"""
    # 检查测图数据中的商品→店铺映射
    data_dir = r"C:\Users\s\Desktop\测图数据"
    try:
        import openpyxl
        for fname in os.listdir(data_dir):
            if not fname.endswith(".xlsx") or fname.startswith("~$"):
                continue
            fpath = os.path.join(data_dir, fname)
            wb = openpyxl.load_workbook(fpath, data_only=True)
            for sn in wb.sheetnames:
                if sn == "Sheet1":
                    continue
                # sheet 名格式: {商品ID}_{日期} 或 {店铺名}_{日期}
                sheet_pid = sn.split("_")[0]
                if sheet_pid == product_id:
                    # 文件名包含店铺名
                    store = fname.replace(".xlsx", "")
                    # 去掉日期后缀
                    import re
                    m = re.match(r"^(.+?)_\d{4}-\d{2}-\d{2}$", store)
                    if m:
                        store = m.group(1)
                    wb.close()
                    return store if store != "_" else "未知店铺"
            wb.close()
    except Exception:
        pass
    return None


async def login_to_pdd(page, store_name, credentials):
    """登录拼多多商家后台，自动填账号密码，暂停等手动过验证码"""
    log("登录店铺: {}".format(store_name))
    cred = credentials.get(store_name)
    if not cred:
        log("  未找到凭据，尝试手动登录...")
        await page.goto(MMS_LOGIN, wait_until="domcontentloaded", timeout=30000)
        log("  请在浏览器中手动登录，完成后按 Enter 继续...")
        input("登录完成后按 Enter: ")
        await page.wait_for_timeout(2000)
        return True

    username = cred["username"]
    password = cred["password"]

    max_retry = 2
    for attempt in range(max_retry):
        log("  登录尝试 {}/{}".format(attempt + 1, max_retry))
        await page.goto(MMS_LOGIN, wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(2000)

        # 检查是否已登录
        if GOODS_LIST in page.url or "goods" in page.url:
            log("  已登录，跳过")
            return True

        # 方式A: 手机号登录 tab
        try:
            phone_tab = await page.query_selector("text=手机号登录")
            if not phone_tab:
                phone_tab = await page.query_selector("[class*='phone'], [class*='Phone']")
            if phone_tab:
                await phone_tab.click()
                await page.wait_for_timeout(500)
        except Exception:
            pass

        # 填账号
        username_selectors = [
            "input[placeholder*='手机号']",
            "input[placeholder*='账号']",
            "input[type='text']",
            "input[name='username']",
            "input[name='mobile']",
        ]
        filled_user = False
        for sel in username_selectors:
            try:
                el = await page.query_selector(sel)
                if el and await el.is_visible():
                    await el.click()
                    await el.fill("")
                    await el.type(username, delay=50)
                    filled_user = True
                    log("  已填账号")
                    break
            except Exception:
                continue

        # 填密码
        pwd_selectors = [
            "input[placeholder*='密码']",
            "input[type='password']",
            "input[name='password']",
        ]
        filled_pwd = False
        for sel in pwd_selectors:
            try:
                el = await page.query_selector(sel)
                if el and await el.is_visible():
                    await el.click()
                    await el.fill("")
                    await el.type(password, delay=50)
                    filled_pwd = True
                    log("  已填密码")
                    break
            except Exception:
                continue

        # 点登录
        try:
            login_btn = await page.query_selector(
                "button:has-text('登录'), [class*='login-btn'], [class*='submit']"
            )
            if login_btn:
                await login_btn.click()
                log("  已点击登录")
        except Exception:
            pass

        # 等待验证码
        log("  ⏳ 请在浏览器中完成滑块验证码（60秒）...")
        await page.wait_for_timeout(3000)

        # 等待跳转到商品列表
        for _ in range(60):
            await page.wait_for_timeout(1000)
            current = page.url
            if GOODS_LIST in current or "/goods" in current:
                log("  登录成功！")
                return True
            # 检查是否还在登录页
            if "login" in current.lower() or "passport" in current.lower():
                continue
            else:
                # 可能跳到了其他页面
                log("  页面已跳转: {}".format(current[:80]))
                if "mms.pinduoduo.com" in current:
                    return True

        log("  登录超时，重试...")

    log("  登录失败")
    return False


async def swap_images(command_json):
    cmd = json.loads(command_json) if isinstance(command_json, str) else command_json
    source = cmd["source"]
    targets = cmd["targets"]

    source_pid = source["product_id"]
    source_img_url = source["image_url"]

    log("=" * 50)
    log("换图任务开始")
    log("源: {} | 图: {}".format(source_pid, source_img_url[:60]))
    log("目标: {} 个商品".format(len(targets)))

    # 1. 下载源图片
    src_filename = sanitize_filename(source_img_url)
    src_path = os.path.join(DOWNLOADS_DIR, "src_{}".format(src_filename))

    try:
        import urllib.request
        urllib.request.urlretrieve(source_img_url, src_path)
        log("源图已下载: {} bytes".format(os.path.getsize(src_path)))
    except Exception as e:
        log("下载源图失败: {}".format(e))
        # 备用：Playwright 内下载
        src_path = None

    # 2. 读取凭据
    creds = get_store_credentials()

    # 3. 确定各目标商品需要登录的店铺
    store_targets = {}  # {store_name: [targets]}
    for t in targets:
        pid = t["product_id"]
        store = find_store_for_product(pid)
        if not store:
            store = "未知店铺"
        store_targets.setdefault(store, []).append(t)

    log("涉及店铺: {}".format(list(store_targets.keys())))

    # 4. 启动 Playwright
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            channel="chrome",
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            viewport={"width": 1400, "height": 900},
        )
        page = await context.new_page()

        results = []

        # 5. 按店铺逐个处理
        for store_name, store_target_list in store_targets.items():
            log("\n--- 店铺: {} ({}个商品) ---".format(store_name, len(store_target_list)))

            # 5a. 登录该店铺
            logged_in = await login_to_pdd(page, store_name, creds)
            if not logged_in:
                log("登录失败，跳过该店铺")
                for t in store_target_list:
                    results.append({
                        "product_id": t["product_id"],
                        "success": False,
                        "error": "店铺登录失败: {}".format(store_name),
                    })
                continue

            # 5b. 如果还没下载源图，先在浏览器中下载
            if not src_path or not os.path.exists(src_path):
                log("在浏览器中获取源图片...")
                await page.goto("{}?goods_id={}".format(GOODS_EDIT, source_pid),
                               wait_until="domcontentloaded", timeout=30000)
                await page.wait_for_timeout(3000)
                # 截图保存
                src_path = os.path.join(DOWNLOADS_DIR,
                    "src_manual_{}.png".format(source_pid))
                await page.screenshot(path=src_path)
                log("截图保存: {}".format(src_path))

            # 5c. 逐个处理目标商品
            for i, target in enumerate(store_target_list):
                target_pid = target["product_id"]
                log("\n[{}/{}] {}".format(i + 1, len(store_target_list), target_pid))

                try:
                    # 打开目标商品编辑页
                    edit_url = "{}?goods_id={}".format(GOODS_EDIT, target_pid)
                    await page.goto(edit_url, wait_until="domcontentloaded", timeout=30000)
                    await page.wait_for_timeout(3000)

                    # 检查是否被踢到登录页
                    if "login" in page.url.lower():
                        log("会话过期，重新登录...")
                        logged_in = await login_to_pdd(page, store_name, creds)
                        if not logged_in:
                            raise Exception("重新登录失败")
                        await page.goto(edit_url, wait_until="domcontentloaded", timeout=30000)
                        await page.wait_for_timeout(3000)

                    # 查找图片上传区域
                    file_inputs = await page.query_selector_all("input[type='file']")
                    log("找到 {} 个文件上传入口".format(len(file_inputs)))

                    if file_inputs:
                        # 替换最差图位置（最后一个或根据 replace_image_type 匹配）
                        replace_idx = -1
                        replace_type = target.get("replace_image_type", "")

                        # 尝试匹配图片类型
                        if replace_type and len(file_inputs) > 1:
                            for j in range(len(file_inputs)):
                                if j == 0 and replace_type == "主轮播图":
                                    replace_idx = 0
                                    break
                                # 副轮播图用后面的位置
                        if replace_idx < 0:
                            replace_idx = len(file_inputs) - 1

                        log("替换位置: 第 {} 个 (共 {} 个)".format(replace_idx + 1, len(file_inputs)))

                        # 上传新图片
                        await file_inputs[replace_idx].set_input_files(src_path)
                        log("已上传图片")
                        await page.wait_for_timeout(2000)

                        # 保存
                        save_selectors = [
                            "button:has-text('保存')",
                            "button:has-text('提交')",
                            "[class*='save']",
                            "[class*='submit']",
                            "button:has-text('发布')",
                        ]
                        saved = False
                        for sel in save_selectors:
                            try:
                                btn = await page.query_selector(sel)
                                if btn and await btn.is_visible():
                                    await btn.click()
                                    log("已点击保存")
                                    saved = True
                                    await page.wait_for_timeout(3000)
                                    break
                            except Exception:
                                continue

                        if not saved:
                            log("未找到保存按钮，页面可能自动保存")

                        results.append({
                            "product_id": target_pid,
                            "success": True,
                            "message": "已替换第 {} 个图位".format(replace_idx + 1),
                        })
                    else:
                        # 截图排查
                        debug_path = os.path.join(DOWNLOADS_DIR,
                            "debug_{}.png".format(target_pid))
                        await page.screenshot(path=debug_path)
                        log("未找到上传入口，截图: {}".format(debug_path))
                        results.append({
                            "product_id": target_pid,
                            "success": False,
                            "error": "未找到图片上传区域",
                        })

                except Exception as e:
                    log("失败: {}".format(e))
                    import traceback
                    traceback.print_exc()
                    results.append({
                        "product_id": target_pid,
                        "success": False,
                        "error": str(e),
                    })

        await browser.close()

    # 6. 输出结果
    log("\n" + "=" * 50)
    ok = sum(1 for r in results if r.get("success"))
    fail = len(results) - ok
    log("完成: {} 成功, {} 失败".format(ok, fail))

    result_file = os.path.join(DOWNLOADS_DIR,
        "swap_result_{}.json".format(datetime.now().strftime("%Y%m%d_%H%M%S")))
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump({"results": results, "source": source_pid,
                   "ok": ok, "fail": fail}, f, ensure_ascii=False, indent=2)
    log("结果: {}".format(result_file))

    return results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--command", required=True, help="JSON 换图命令")
    args = parser.parse_args()
    asyncio.run(swap_images(args.command))


if __name__ == "__main__":
    main()
